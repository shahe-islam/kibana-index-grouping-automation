import openpyxl

#example open_path /Users/Shahe.Islam/developer/ndap-journey/ndap-journey.xlsx

open_path = input("Input the file open path: ")


wb = openpyxl.load_workbook(open_path)
ws = wb['Sheet 1']

###################
##### SIZES #######
###################

LOG_SIZE_GB = 'gb'
LOG_SIZE_MB = 'mb'

###################
##### JOURNEY #####
###################

# journeys = {
#     EDB = ['edb', 'containerlogs-edb'],
#     HOME = ['home', 'containerlogs-home'],
#     OB = ['ob', 'obcompete'],
#     MYNW = ['mynw', 'mynationwide'],
#     RAAS = ['raas'],
#     NDAP = ['ndap', 'ops', 'containerlogs', 'telegraf', 'beat', 'tgw', 'watcher', 'prod', 'monitoring'],
# }

EDB = ['edb', 'containerlogs-edb']
HOME = ['home', 'containerlogs-home']
OB = ['ob', 'obcompete']
MYNW = ['mynw', 'mynationwide']
RAAS = ['raas']
NDAP = ['ndap', 'ops', 'containerlogs', 'telegraf', 'beat', 'tgw', 'watcher', 'prod', 'monitoring']

for i in range(2, ws.max_row + 1):

    size = ws.cell(row=i, column=9).value

    if LOG_SIZE_GB in size:
        size = size.strip(LOG_SIZE_GB)
        size = str(float(size)*10000)
    elif LOG_SIZE_MB in size:
        size = size.strip(LOG_SIZE_MB)
    else:
        size = 0

    ws.cell(row=i, column=11).value = size

    index = ws.cell(row=i, column=3).value

    if any(x in index for x in EDB):
        journey = 'EDB'
    elif any(x in index for x in HOME):
        journey = 'HOME'
    elif any(x in index for x in OB):
        journey = 'OB'
    elif any(x in index for x in MYNW):
        journey = 'MYNW'
    elif any(x in index for x in RAAS):
        journey = 'RAAS'
    elif any(x in index for x in NDAP):
        journey = 'NDAP'
    else:
        journey = 'default'
    
    ws.cell(row=i, column=12).value = journey
    journey = ''


#example save_path /Users/Shahe.Islam/developer/ndap-journey/ndap-journey-test2.xlsx

save_path = input("Input the file save path: ")
wb.save(save_path)