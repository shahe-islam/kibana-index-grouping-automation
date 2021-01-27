import openpyxl

###################
##### COLUMNS #####
###################

INDEX = 3
ORIGINAL_SIZE = 9
SCALED_SIZE = 11
HUB = 12

###################
##### SIZES #######
###################

INDEX_SIZE_GB = 'gb'
INDEX_SIZE_MB = 'mb'

###################
###### HUBS #######
###################

hubs = {
    'EDB':['edb', 'containerlogs-edb'],
    'HOME':['home', 'containerlogs-home'],
    'OB':['ob', 'obcompete'],
    'MYNW':['mynw', 'mynationwide'],
    'RAAS':['raas'],
    'NDAP':['ndap', 'ops', 'containerlogs', 'telegraf', 'beat', 'tgw', 'watcher', 'prod', 'monitoring'],
}

#/Users/Shahe.Islam/developer/ndap-journey/ndap-journey.xlsx

open_path = input("Input the file open path: ")

wb = openpyxl.load_workbook(open_path)
ws = wb['Sheet 1']

for k in hubs:
    wb.create_sheet(k)
    nws = wb[k]

    for i,row in enumerate(ws.iter_rows(max_row=1)):
        for j,col in enumerate(row):
            nws.cell(row=i+1,column=j+1).value = col.value

def index_size_scaler(size):
    if INDEX_SIZE_GB in size:
        size = size.strip(INDEX_SIZE_GB)
    elif INDEX_SIZE_MB in size:
        size = size.strip(INDEX_SIZE_MB)
        size = float(size)/1024
    else:
        size = 0
    return size

for i in range(2, ws.max_row + 1):

    size = ws.cell(row=i, column=ORIGINAL_SIZE).value
    scaled_size = index_size_scaler(size)
    ws.cell(row=i, column=SCALED_SIZE).value = float(scaled_size)

    index = ws.cell(row=i, column=INDEX).value

    for hub, journey in hubs.items():
        if any(x in index for x in journey): break

        ##TODO Add functionality to automatically copy current row in loop to row in matching worksheet

    else: hub = 'default'

    ws.cell(row=i, column=HUB).value = hub

#/Users/Shahe.Islam/developer/ndap-journey/ndap-journey-test.xlsx

save_path = input("Input the file save path: ")
wb.save(save_path)
print("Your file has been saved at: " + save_path)